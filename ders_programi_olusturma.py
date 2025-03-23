from sqlalchemy.orm import sessionmaker, declarative_base, relationship
import random
from sqlalchemy import create_engine, inspect, Column, Integer, String, ForeignKey, text,  Table, select
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter

# **Veritabanı Bağlantı Bilgileri**
DB_NAME = "DersProgramiDB"
SERVER_NAME = "DESKTOP-0P4T7M7"  # MSSQL sunucu adı
DRIVER = "ODBC+Driver+17+for+SQL+Server"

# **Bağlantı URL'leri**
DB_URL = f"mssql+pyodbc://{SERVER_NAME}/{DB_NAME}?trusted_connection=yes&driver={DRIVER}"
MASTER_DB_URL = f"mssql+pyodbc://{SERVER_NAME}/master?trusted_connection=yes&driver={DRIVER}"

# **SQLAlchemy Motoru ve Base Tanımlama**
engine = create_engine(DB_URL)
Base = declarative_base()


# **Veritabanı Var mı Kontrol Et**
def database_exists():
    try:
        engine_temp = create_engine(MASTER_DB_URL)
        with engine_temp.connect() as conn:
            result = conn.execute(text(f"SELECT name FROM sys.databases WHERE name = '{DB_NAME}'"))
            return result.fetchone() is not None
    except Exception as e:
        print(f"⚠️ Veritabanı kontrol edilirken hata oluştu: {e}")
        return False


# **Veritabanı Oluşturma (AUTOCOMMIT ile)**
def create_database():
    try:
        engine_temp = create_engine(MASTER_DB_URL, isolation_level="AUTOCOMMIT")  # AUTOCOMMIT etkin
        with engine_temp.connect() as conn:
            conn.execute(text(f"CREATE DATABASE {DB_NAME}"))
        print(f"✅ Veritabanı '{DB_NAME}' başarıyla oluşturuldu.")
    except Exception as e:
        print(f"❌ Veritabanı oluşturulurken hata oluştu: {e}")


# **Eğer Veritabanı Yoksa Oluştur**
if not database_exists():
    create_database()

# **MODELLER**
# Öğretim Üyesi - Bölüm ilişki tablosu
ogretim_uyesi_bolum = Table(
    "ogretim_uyesi_bolum", Base.metadata,
    Column("ogretim_uyesi_id", Integer, ForeignKey("kullanicilar.id")),
    Column("bolum_kod", String(10), ForeignKey("bolum.kod")),
)


# **Kullanıcılar Tablosu**
class Kullanicilar(Base):
    __tablename__ = "kullanicilar"
    id = Column(Integer, primary_key=True)
    mevki = Column(String(20), nullable=False)  # "OGRETIM_UYESI", "OGRENCI", "YONETICI"
    ad = Column(String(100), nullable=False)

    # Öğretim üyesinin dersleri (eğer öğretim üyesi ise)
    dersler = relationship("Ders", back_populates="ogretim_uyesi")

    # Öğretim üyesinin bağlı olduğu bölümler
    bolumler = relationship("Bolum", secondary=ogretim_uyesi_bolum, back_populates="ogretim_uyeleri")


# 🏛 **Bölüm Tablosu**
class Bolum(Base):
    __tablename__ = "bolum"
    id = Column(Integer, primary_key=True)  # Bölüm ID'si
    kod = Column(String(10), unique=True, nullable=False)
    ad = Column(String(100), nullable=False)

    # Bölümdeki dersler (bolum_id ForeignKey ile ilişkilendirilmeli)
    dersler = relationship("Ders", back_populates="bolum")

    # Bölüme bağlı öğretim üyeleri (Many-to-Many ilişkisi)
    ogretim_uyeleri = relationship("Kullanicilar", secondary=ogretim_uyesi_bolum, back_populates="bolumler")


# 📚 **Ders Tablosu**
class Ders(Base):
    __tablename__ = "ders"
    id = Column(Integer, primary_key=True)
    bolum_kod = Column(String(10), ForeignKey("bolum.kod"))  # ForeignKey, Bolum.kod ile ilişkilendirildi
    donem = Column(Integer, nullable=False)
    kod = Column(String(20), nullable=False)
    ad = Column(String(100), nullable=False)
    ders_tipi = Column(String(50), nullable=False)
    teorik_saat = Column(Integer, nullable=False)
    uyg_saat = Column(Integer, nullable=False)
    ogretim_uyesi_id = Column(Integer, ForeignKey("kullanicilar.id"))  # Öğretim üyesi (Kullanıcılar tablosu)

    # Relationships
    bolum = relationship("Bolum", back_populates="dersler")  # Dersin bağlı olduğu bölüm
    ogretim_uyesi = relationship("Kullanicilar", back_populates="dersler")


# 🏢 **Derslik Tablosu**
class Derslik(Base):
    __tablename__ = "derslik"
    id = Column(Integer, primary_key=True)
    kod = Column(String(20), unique=True, nullable=False)
    kapasite = Column(Integer, nullable=False)
    statu = Column(String(20), nullable=False)  # NORMAL / LAB


# **Veritabanındaki Tabloları Oluştur**
Base.metadata.create_all(engine)


# **Tablolar Var mı Kontrol Et**
def tables_exist():
    try:
        inspector = inspect(engine)
        required_tables = {"bolum", "ders", "derslik", "ogretim_uyesi"}
        existing_tables = set(inspector.get_table_names())
        return required_tables.issubset(existing_tables)
    except Exception as e:
        print(f"⚠️ Tablolar kontrol edilirken hata oluştu: {e}")
        return False


# **Eğer Tablolar Yoksa Oluştur**
if not tables_exist():
    try:
        Base.metadata.create_all(engine)
        print("✅ Tablolar başarıyla oluşturuldu.")
    except Exception as e:
        print(f"❌ Tablolar oluşturulurken hata oluştu: {e}")

# **Session Oluştur (Veritabanı İşlemleri İçin)**
try:
    Session = sessionmaker(bind=engine)
    session = Session()
    print("✅ Veritabanı bağlantısı başarıyla kuruldu.")
except Exception as e:
    print(f"❌ Veritabanı bağlantısı kurulurken hata oluştu: {e}")


def load_kullanicilar_from_file(filename):
    """kullanicilar.txt dosyasından kullanıcıları oku ve veritabanına ekle"""
    try:
        with open(filename, "r", encoding="utf-8-sig") as file:
            yeni_kullanici_eklendi = False

            for line in file:
                data = line.strip().split(maxsplit=2)  # İlk iki değeri al, ad kısmı boşluk içerebilir

                if len(data) != 3:
                    print(f"⚠️ Hatalı format: {line.strip()}")
                    continue

                kullanici_id, mevki, ad = data
                kullanici_id = int(kullanici_id)

                # Eğer kullanıcı daha önce eklenmemişse, ekle
                if not session.query(Kullanicilar).filter_by(id=kullanici_id).first():
                    session.add(Kullanicilar(id=kullanici_id, ad=ad, mevki=mevki))
                    yeni_kullanici_eklendi = True

        if yeni_kullanici_eklendi:
            session.commit()
            print("✅ Kullanıcılar başarıyla eklendi!")
        else:
            print("✅ Kullanıcılar zaten mevcut.")

    except FileNotFoundError:
        print("❌ Hata: kullanicilar.txt dosyası bulunamadı!")
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")

def load_ogretim_uyesi_bolum_from_file(filename):
    """ogretim_uyesi_bolum.txt dosyasından ilişkiyi oku ve veritabanına ekle"""
    try:
        with open(filename, "r", encoding="utf-8-sig") as file:
            yeni_iliski_eklendi = False

            for line in file:
                data = line.strip().split()

                if len(data) != 2:
                    print(f"⚠️ Hatalı format: {line.strip()}")
                    continue

                ogretim_uyesi_id, bolum_kod = data
                ogretim_uyesi_id = int(ogretim_uyesi_id)

                # öğretim üyesi ile bölüm arasındaki ilişkinin zaten veritabanında olup olmadığını kontrol eder.
                stmt = select(ogretim_uyesi_bolum).where(
                    (ogretim_uyesi_bolum.c.ogretim_uyesi_id == ogretim_uyesi_id) &
                    (ogretim_uyesi_bolum.c.bolum_kod == bolum_kod)
                )
                existing_relation = session.execute(stmt).first()

                if not existing_relation:
                    session.execute(ogretim_uyesi_bolum.insert().values(
                        ogretim_uyesi_id=ogretim_uyesi_id, bolum_kod=bolum_kod
                    ))
                    yeni_iliski_eklendi = True

        if yeni_iliski_eklendi:
            session.commit()
            print("✅ Öğretim Üyesi - Bölüm ilişkileri başarıyla eklendi!")
        else:
            print("✅ Tüm ilişkiler zaten mevcut.")

    except FileNotFoundError:
        print("❌ Hata: ogretim_uyesi_bolum.txt dosyası bulunamadı!")
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")


def load_derslikler_from_file(filename):
    """derslik.txt dosyasından derslikleri oku ve veritabanına ekle"""
    try:
        with open(filename, "r", encoding="utf-8-sig") as file:
            derslik_eklendi = False  # Yeni derslik eklenip eklenmediğini takip etmek için

            for line in file:
                data = line.strip().split()  # Satırı boşluklara göre ayır

                # Verinin doğru formatta olup olmadığını kontrol et
                if len(data) != 3:
                    print(f"⚠️ Hatalı format: {line.strip()}")
                    continue

                kod, kapasite, statu = data
                kapasite = int(kapasite)  # Kapasiteyi integer'a çevir

                # Eğer bu derslik daha önce eklenmemişse, ekle
                if not session.query(Derslik).filter_by(kod=kod).first():
                    session.add(Derslik(kod=kod, kapasite=kapasite, statu=statu))
                    derslik_eklendi = True  # Yeni derslik eklendiğini işaretle

        if derslik_eklendi:
            session.commit()
            print("✅ Derslikler başarıyla eklendi!")
        else:
            print("✅ Derslikler zaten mevcut.")  # Hiçbir yeni derslik eklenmediyse mesaj ver

    except FileNotFoundError:
        print("❌ Hata: derslik.txt dosyası bulunamadı!")
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")


def load_dersler_from_file(filename):
    try:
        # Bölüm tablosunun boş olup olmadığını kontrol et
        bolum_count = session.query(Bolum).count()
        if bolum_count == 0:
            print("❌ Lütfen önce Bölüm tanımlayınız.")
            return  # Bölüm tablosu boşsa fonksiyon sonlanır

        with open(filename, "r", encoding="utf-8-sig") as file:
            ders_eklendi = False  # Yeni ders eklenip eklenmediğini takip etmek için

            for line in file:
                data = line.strip().split()  # Satırı boşluklara göre ayır

                # Verinin doğru formatta olup olmadığını kontrol et
                if len(data) != 7:
                    print(f"⚠️ Hatalı format: {line.strip()}")
                    continue

                bolum_kod, donem, kod, ad, ders_tipi, teorik_saat, uyg_saat = data
                donem = int(donem)
                teorik_saat = int(teorik_saat)
                uyg_saat = int(uyg_saat)

                # Bolum koduna göre ilgili bolum nesnesini al
                bolum_instance = session.query(Bolum).filter_by(kod=bolum_kod).first()

                if not bolum_instance:
                    print(f"⚠ Bölüm bulunamadı: {bolum_kod}")
                    continue  # Eğer bölüm bulunmazsa, bu satırda işlem yapma

                benzersiz_ders_kodu = f"{bolum_kod}_{kod}"  # BM_MAT110, YM_MAT110 gibi

                # Eğer bu ders daha önce eklenmemişse, ekle
                if not session.query(Ders).filter_by(kod=benzersiz_ders_kodu).first():
                    new_ders = Ders(
                        bolum=bolum_instance,  # Burada bolum nesnesini ilişkilendiriyoruz
                        donem=donem,
                        kod=benzersiz_ders_kodu,
                        ad=ad,
                        ders_tipi=ders_tipi,
                        teorik_saat=teorik_saat,
                        uyg_saat=uyg_saat
                    )
                    session.add(new_ders)
                    ders_eklendi = True  # Yeni ders eklendiğini işaretle

        if ders_eklendi:
            session.commit()
            print("✅ Dersler başarıyla eklendi!")
        else:
            print("✅ Dersler zaten mevcut.")

    except FileNotFoundError:
        print("❌ Hata: dersler.txt dosyası bulunamadı!")
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")


# **Öğretim üyelerine rastgele 3 farklı ders atama ve ortak derslere aynı öğretim üyesini atama fonksiyonu**
def assign_random_courses():
    try:
        ogretim_uyeleri = session.query(Kullanicilar).filter_by(mevki="ogretim_uyesi").all()
        dersler = session.query(Ders).filter(Ders.ogretim_uyesi_id == None).all()
        ortak_dersler = defaultdict(list)
        
        for ders in dersler:
            ortak_dersler[ders.ad].append(ders)  # Aynı ada sahip dersleri grupla
        
        random.shuffle(ogretim_uyeleri)  # Rastgele sırayla öğretmenleri işle
        ogretim_ders_sayisi = {hoca.id: 0 for hoca in ogretim_uyeleri}  # Her hocaya 3 farklı ders sınırı koy
        
        for ders_ad, ders_listesi in ortak_dersler.items():
            uygun_hocalar = [hoca for hoca in ogretim_uyeleri if ogretim_ders_sayisi[hoca.id] <= 3]
            
            if not uygun_hocalar:
                print(f"⚠ {ders_ad} için uygun öğretim üyesi bulunamadı!")
                continue

            secilen_hoca = random.choice(uygun_hocalar)  # Rastgele bir öğretim üyesi seç

            for ders in ders_listesi:
                ders.ogretim_uyesi_id = secilen_hoca.id  # Tüm ortak derslere aynı hocayı ata
                print(f"✅ {secilen_hoca.ad} → {ders.ad} ({ders.bolum_kod}) dersine atandı.")
                
            ogretim_ders_sayisi[secilen_hoca.id] += 1  # Ortak dersleri tek ders olarak say
            
            if ogretim_ders_sayisi[secilen_hoca.id] >= 3 and secilen_hoca in ogretim_uyeleri:
                ogretim_uyeleri.remove(secilen_hoca)  # 3 farklı ders dolunca listeden çıkar (Hata önlendi)

        session.commit()
        print("🎉 Ders atamaları başarıyla tamamlandı!")
    except Exception as e:
        print(f"❌ Beklenmeyen hata: {e}")



class Sistem:
    def menu():
        while True:
            print("\n||---------- ANA MENU ----------||\n",
                  "  1. Bölüm İşlemleri\n"
                  "  2. Kullanıcı İşlemleri\n",
                  "  3. Ders İşlemleri\n",
                  "  4. Derslik İşlemleri\n",
                  "  5. Ders Programı Oluştur\n",
                  "  6. Çıkış")

            secim = input("Seçiminizi yapın (1-6): ")

            if secim == "1":
                Sistem.bolum_islemleri()
            elif secim == "2":
                Sistem.kullanici_islemleri()
            elif secim == "3":
                Sistem.ders_islemleri()
            elif secim == "4":
                Sistem.derslik_islemleri()
            elif secim == "5":
                Sistem.ders_programi_olustur()
            elif secim == "6":
                print("Çıkış yapılıyor...")
                break
            else:
                print("Hatalı giriş! Lütfen 1-6 arasında bir değer girin.")

    def kullanici_islemleri():
        while True:
            print("\n||------ KULLANICI İŞLEMLERİ ------||\n",
                  "  1. Kullanıcıları Dosyadan Oku\n",
                  "  2. Kullanıcı Ekle\n",
                  "  3. Kullanıcı Sil\n",
                  "  4. Kullanıcı Düzenle\n",
                  "  5. Geri Dön")

            secim = input("Seçiminizi yapın (1-5): ")
            if secim == "1":
                ad = input("Txt Dosyasının Adı (kullanicilar): \n")
                bol_ad = input(" Txt Dosyasının Adı (ogretim_uyesi_bolum): \n")
                load_kullanicilar_from_file(f"{ad}.txt")
                load_ogretim_uyesi_bolum_from_file(f"{bol_ad}.txt")
            elif secim == "2":
                Sistem.kullanici_ekle()
            elif secim == "3":
                Sistem.kullanici_sil()
            elif secim == "4":
                Sistem.kullanici_duzenle()
            elif secim == "5":
                print("Ana menüye dönülüyor...\n")
                break
            else:
                print("Hatalı giriş! Lütfen 1-5 arasında bir değer girin.")

    def kullanici_ekle():
        print("\n||--- Yeni Kullanıcı Ekle ---||")
        ad = input("Kullanıcının adı: ").strip()
        mevki = input("Mevki (öğretim_üyesi, öğrenci, yönetici): ").strip()

        if mevki not in ["öğretim_üyesi", "öğrenci", "yönetici"]:
            print("Hatalı mevki girdiniz! Lütfen doğru bir değer girin.")
            return

        yeni_kullanici = Kullanicilar(ad=ad, mevki=mevki)
        session.add(yeni_kullanici)
        session.commit()  # Kullanıcı ID'sinin oluşması için commit gerekiyor

        if mevki == "öğretim_üyesi":
            bolum_kodlari = input("Öğretim üyesinin ait olduğu bölümlerin kodlarını (virgülle ayırarak) girin: ").split(
                ',')
            bolumler = session.query(Bolum).filter(Bolum.kod.in_([b.strip() for b in bolum_kodlari])).all()

            if not bolumler:
                print("❌ Hata: Girilen bölüm kodlarına ait bir kayıt bulunamadı.")
                return

            # `ogretim_uyesi_bolum` ilişki tablosuna doğrudan ekleme yap
            for bolum in bolumler:
                session.execute(
                    ogretim_uyesi_bolum.insert().values(
                        ogretim_uyesi_id=yeni_kullanici.id,
                        bolum_kod=bolum.kod
                    )
                )

            session.commit()  # Değişiklikleri kaydet
            print(f"✅ Öğretim üyesi {yeni_kullanici.ad} aşağıdaki bölümlere atanmıştır: {[b.ad for b in bolumler]}")

        print(f"\n✅ Kullanıcı eklendi: {yeni_kullanici.ad} - {yeni_kullanici.mevki} ve ilişkili bölümler kaydedildi.")

    def kullanici_sil():
        print("\n||--- Kullanıcı Sil ---||")
        kullanici_ad = input("Silmek istediğiniz kullanıcının adını girin: ").strip()
        kullanici = session.query(Kullanicilar).filter_by(ad=kullanici_ad).first()

        if kullanici:
            session.delete(kullanici)  # Otomatik olarak ogretim_uyesi_bolum'dan da siler (CASCADE sayesinde)
            session.commit()
            print(f"\n✅ Kullanıcı silindi: {kullanici.ad} ve ilişkili bölümler kaldırıldı.")
        else:
            print("❌ Kullanıcı bulunamadı!")

    def kullanici_duzenle():
        print("\n||--- Kullanıcı Özelliği Düzenle ---||")
        kullanici_ad = input("Düzenlemek istediğiniz kullanıcının adını girin: ").strip()
        kullanici = session.query(Kullanicilar).filter_by(ad=kullanici_ad).first()

        if not kullanici:
            print("❌ Hata: Böyle bir kullanıcı bulunamadı.")
            return

        ozellik = input("Düzenlemek istediğiniz kullanıcı özelliğini seçiniz (Mevki, Bölüm): ").strip().lower()

        if ozellik == "mevki":
            yeni_mevki = input("Yeni mevkiyi girin (öğretim_üyesi, öğrenci, yönetici): ").strip()

            if yeni_mevki not in ["öğretim_üyesi", "öğrenci", "yönetici"]:
                print("❌ Hata: Geçersiz mevki girdiniz.")
                return

            kullanici.mevki = yeni_mevki
            session.commit()
            print(f"✅ Kullanıcının mevki bilgisi güncellendi: {kullanici.ad} - {kullanici.mevki}")

        elif ozellik == "bölüm":
            if kullanici.mevki != "öğretim_üyesi":
                print("❌ Hata: Sadece öğretim üyelerinin bölümü değiştirilebilir.")
                return

            mevcut_bolumler = session.query(Bolum).join(ogretim_uyesi_bolum).filter(
                ogretim_uyesi_bolum.c.ogretim_uyesi_id == kullanici.id
            ).all()

            print(f"📌 Kullanıcının mevcut bölümleri: {[b.ad for b in mevcut_bolumler]}")

            islem = input("Bölüm eklemek mi (E) yoksa çıkarmak mı (Ç) istiyorsunuz? ").strip().lower()

            if islem == "e":
                yeni_bolum_kodlari = input("Eklemek istediğiniz bölüm kodlarını (virgülle ayırarak) girin: ").split(',')
                yeni_bolumler = session.query(Bolum).filter(
                    Bolum.kod.in_([b.strip() for b in yeni_bolum_kodlari])).all()

                if not yeni_bolumler:
                    print("❌ Hata: Girilen bölüm kodlarına ait bir kayıt bulunamadı.")
                    return

                for bolum in yeni_bolumler:
                    session.execute(
                        ogretim_uyesi_bolum.insert().values(
                            ogretim_uyesi_id=kullanici.id,
                            bolum_kod=bolum.kod
                        )
                    )

                session.commit()
                print(f"✅ Kullanıcının bölümleri güncellendi: {[b.ad for b in yeni_bolumler]} eklendi.")

            elif islem == "ç":
                silinecek_bolum_kodlari = input("Silmek istediğiniz bölüm kodlarını (virgülle ayırarak) girin: ").split(
                    ',')
                silinecek_bolumler = session.query(Bolum).filter(
                    Bolum.kod.in_([b.strip() for b in silinecek_bolum_kodlari])).all()

                if not silinecek_bolumler:
                    print("❌ Hata: Girilen bölüm kodlarına ait bir kayıt bulunamadı.")
                    return

                for bolum in silinecek_bolumler:
                    session.execute(
                        ogretim_uyesi_bolum.delete().where(
                            (ogretim_uyesi_bolum.c.ogretim_uyesi_id == kullanici.id) &
                            (ogretim_uyesi_bolum.c.bolum_kod == bolum.kod)
                        )
                    )

                session.commit()
                print(f"✅ Kullanıcının bölümleri güncellendi: {[b.ad for b in silinecek_bolumler]} çıkarıldı.")

            else:
                print("❌ Hata: Geçersiz seçenek girdiniz.")

        else:
            print("❌ Hata: Lütfen sadece 'Mevki' veya 'Bölüm' giriniz.")

    def bolum_islemleri():
        while True:
            print("\n||------ BÖLÜM İŞLEMLERİ ------||\n",
                  "  1. Bölüm Ekle\n",
                  "  2. Bölüm Sil\n",
                  "  3. Geri Dön")

            secim = input("Seçiminizi yapın (1-3): ")

            if secim == "1":
                Sistem.bolum_ekle()
            elif secim == "2":
                Sistem.bolum_sil()
            elif secim == "3":
                print("Ana menüye dönülüyor...\n")
                break
            else:
                print("Hatalı giriş! Lütfen 1-3 arasında bir değer girin.")

    def bolum_ekle():
        print("\n||--- Yeni Bölüm Ekle ---||")
        ad = input("Bölüm adı: ").strip()
        kod = input("Bölüm Kodu (BM, YM): ").strip()

        yeni_bolum = Bolum(ad=ad, kod=kod)
        session.add(yeni_bolum)
        session.commit()
        print(f"\n✅ Bölüm tanımlandı: {yeni_bolum.ad} - {yeni_bolum.kod}")

    def bolum_sil():
        print("\n||--- Bölüm Sil ---||")
        bolum_kod = input("Silmek istediğiniz Bölüm Kodunu girin: ").strip()
        bolum = session.query(Bolum).filter_by(kod=bolum_kod).first()

        if bolum:
            session.delete(bolum)
            session.commit()
            print(f"\n✅ {bolum.ad} Bölümü silindi.")
        else:
            print("❌ Bölüm bulunamadı!")

    def ders_islemleri():
        while True:
            print("\n||------ DERS İŞLEMLERİ ------||\n",
                  "  1. Dersleri Dosyadan Oku\n",
                  "  2. Derslere Rastgele Öğretmen Ata"
                  "  3. Ders Ekle\n",
                  "  4. Ders Sil\n",
                  "  5. Geri Dön")

            secim = input("Seçiminizi yapın (1-5): ")
            if secim == "1":
                ad = input("Txt Dosyasının Adı: ")
                load_dersler_from_file(f"{ad}.txt")
            elif secim == "2":
                assign_random_courses()
            elif secim == "3":
                Sistem.ders_ekle()
            elif secim == "4":
                Sistem.ders_sil()
            elif secim == "5":
                print("Ana menüye dönülüyor...\n")
                break
            else:
                print("Hatalı giriş! Lütfen 1-5 arasında bir değer girin.")

    def ders_ekle():
        print("\n||--- Yeni Ders Ekle ---||")
        bolum_kod = input("Bölüm Kodu: ").strip()

        bolum = session.query(Bolum).filter_by(kod=bolum_kod).first()

        if not bolum:
            print("❌ Hata: Girilen bölüm koduna sahip bir bölüm bulunamadı.")
            return

        ogretim_uyeleri = session.query(Kullanicilar).join(ogretim_uyesi_bolum).filter(
            ogretim_uyesi_bolum.c.bolum_kod == bolum.kod).all()
        ogretim_listesi = ", ".join([f"{uy.id} - {uy.ad}" for uy in ogretim_uyeleri])

        donem = input("Dersin Dönemi: ").strip()
        kod = input("Ders Kodu: ").strip()
        ad = input("Dersin Adı: ").strip()
        ders_tipi = input("Ders Tipi: ").strip()
        teorik_saat = input("Dersin Teorik Saati: ").strip()
        uyg_saat = input("Dersin Uygulama Saati: ").strip()
        print(f"📌 Bu bölüme ait öğretim üyeleri: {ogretim_listesi}")
        ogretim_uyesi_id = input("Dersin Öğretim Görevlisi ID: ").strip()

        ogretim_uyesi = session.query(Kullanicilar).filter_by(id=ogretim_uyesi_id, mevki="öğretim_üyesi").first()

        if not ogretim_uyesi:
            print("❌ Hata: Girilen ID'ye sahip bir öğretim üyesi bulunamadı.")
            return

        yeni_ders = Ders(bolum_kod=bolum_kod, donem=donem, kod=kod, ad=ad, ders_tipi=ders_tipi, teorik_saat=teorik_saat,
                         uyg_saat=uyg_saat, ogretim_uyesi_id=ogretim_uyesi.id)
        session.add(yeni_ders)
        session.commit()
        print(f"\n✅ Ders tanımlandı: {yeni_ders.kod}")

    def ders_sil():
        print("\n||--- Ders Sil ---||")
        ders_kod = input("Silmek istediğiniz Dersin Kodunu girin: ").strip()
        ders = session.query(Ders).filter_by(kod=ders_kod).first()

        if not ders:
            print("❌ Hata: Girilen ders koduna sahip bir ders bulunamadı.")
            return

        if ders:
            session.delete(ders)
            session.commit()
            print(f"\n✅ {ders.kod} Dersi silindi.")
        else:
            print("❌ Ders bulunamadı!")

    def derslik_islemleri():
        while True:
            print("\n||------ DERSLİK İŞLEMLERİ ------||\n",
                  "  1. Derslikleri Dosyadan Oku\n",
                  "  2. Derslik Ekle\n",
                  "  3. Derslik Sil\n",
                  "  4. Geri Dön")

            secim = input("Seçiminizi yapın (1-4): ")
            if secim == "1":
                ad = input("Txt Dosyasının Adı: ")
                load_derslikler_from_file(f"{ad}.txt")
            elif secim == "2":
                Sistem.derslik_ekle()
            elif secim == "3":
                Sistem.derslik_sil()
            elif secim == "4":
                print("Ana menüye dönülüyor...\n")
                break
            else:
                print("Hatalı giriş! Lütfen 1-4 arasında bir değer girin.")

    def derslik_ekle():
        print("\n||--- Yeni Derslik Ekle ---||")
        kod = input("Derslik Kodu: ").strip()
        kapasite = input("Derslik kapasitesi: ").strip()
        statu = input("Derslik statusu: ").strip()

        yeni_derslik = Derslik(kod=kod, kapasite=kapasite, statu=statu)
        session.add(yeni_derslik)
        session.commit()
        print(f"\n✅ Derslik tanımlandı: {yeni_derslik.kod}")

    def derslik_sil():
        print("\n||--- Derslik Sil ---||")
        derslik_kod = input("Silmek istediğiniz Derslik Kodunu girin: ").strip()
        derslik = session.query(Derslik).filter_by(kod=derslik_kod).first()

        if derslik:
            session.delete(derslik)
            session.commit()
            print(f"\n✅ {derslik.kod} Dersliği silindi.")
        else:
            print("❌ Derslik bulunamadı!")

    def ders_programi_olustur():

        donem_tipi = input("Güz mü Bahar mı? (G/B): ").strip().upper()
        derslikler = Program.derslikleri_oku()

        if donem_tipi == "G":
            donemler = [1, 3, 5, 7]
            for donem in donemler:
                # Dersleri veritabanından al ve programı oluştur
                Program.dersleri_oku(donem, derslikler)

                print(f"\n✅ {donem}. Dönem için ders programı oluşturuldu.")

        elif donem_tipi == "B":
            donemler = [2, 4, 6, 8]
            for donem in donemler:
                # Dersleri veritabanından al ve programı oluştur
                Program.dersleri_oku(donem, derslikler)

                print(f"\n✅ {donem}. Dönem için ders programı oluşturuldu.")


class Program:

    # ---DERSLİK ATAMA İŞLEMLERİ---
    @staticmethod
    def derslikleri_oku():
        derslikler = []  # Derslikleri saklamak için liste
        """ Veritabanından tüm derslikleri çekip listeye ekler. """
        try:
            derslikler = session.query(Derslik).all()
            return derslikler
        except Exception as e:
            print(f"Derslikleri okuma hatası: {e}")

    @staticmethod
    def uygun_derslik_bul(ders, ortak_dersler, derslikler, kullanilan_derslik, lab_dersi):
        """ Dersin türüne göre uygun derslik seçimi yapar. """
        ortak_d = {ders_bilgi["Ders Adı"] for ders_bilgi in ortak_dersler.values()}
        uygun_derslikler = []

        # Eğer ders bir laboratuvar dersi ise sadece laboratuvar derslikleri seçilecek
        if lab_dersi or "lab" in ders["Ders Adı"].lower():
            uygun_derslikler = [d for d in derslikler if "lab" in d.kod.lower() and d.kod not in kullanilan_derslik]

        # Ortak dersler ve seçmeli dersler için kapasitesi 70'ten büyük olanlar seçilecek
        elif ders["Ders Adı"] in ortak_d or ders["Ders Tipi"].lower() == "secmeli":
            uygun_derslikler = [d for d in derslikler if d.kapasite > 90 and d.kod not in kullanilan_derslik]

        # Diğer dersler için alfabetik olarak sıralanmış uygun derslikler seçilecek
        else:
            uygun_derslikler = sorted(
                [d for d in derslikler if d.kod not in kullanilan_derslik and d.kapasite <80],
                key=lambda d: d.kod
            )

        return uygun_derslikler if uygun_derslikler else derslikler


    @staticmethod
    def derslikleri_atama(bm_program, ym_program, bm_dersler, ym_dersler, bm_ortak_dersler, derslikler):
        """ Programdaki tüm derslere uygun derslik ataması yapar. """
        derslik_doluluk = defaultdict(lambda: defaultdict(set))  # {gun: {saat: {derslik1, derslik2, ...}}}
        ortak_derslik_atama = {}  # Ortak dersler için atanmış derslikleri saklar
        derslik_atama = {}  # Aynı bölümdeki dersler için tek derslik kullanımı
        kullanilan_derslik = set()  # Kullanılan derslikleri takip etmek için

        ortak_d = {ders_bilgi["Ders Adı"] for ders_kodu, ders_bilgi in bm_ortak_dersler.items()}

        for bolum, program in [("BM", bm_program), ("YM", ym_program)]:
            for gun, saatler in program.items():
                for saat, ders in saatler.items():
                    if ders and ders != "":  # Boş olmayan saatler için derslik atayalım
                        ders_kodu, ders_adi = ders

                        # Ders bilgisi al
                        ders_bilgisi = bm_dersler.get(ders_kodu, ym_dersler.get(ders_kodu, None))
                        if not ders_bilgisi:
                            continue

                        # Eğer ortak bir dersse ve daha önce atanmışsa, aynı dersliği kullan
                        if ders_adi in ortak_d and ders_adi in ortak_derslik_atama:
                            program[gun][saat] = (ders_kodu, ders_adi, ortak_derslik_atama[ders_adi])
                            continue  

                        # Aynı bölümdeki aynı ders için daha önce derslik atanmışsa, onu kullan
                        if (bolum, ders_adi) in derslik_atama:
                            program[gun][saat] = (ders_kodu, ders_adi, derslik_atama[(bolum, ders_adi)])
                            continue

                        # Ders bir LAB dersi mi?
                        lab_dersi = "LAB" in ders_kodu.upper()

                        # Uygun derslikleri belirle
                        uygun_derslikler = Program.uygun_derslik_bul(ders_bilgisi, bm_ortak_dersler, derslikler, kullanilan_derslik, lab_dersi)

                        # Ders için uygun bir derslik bul ve ata
                        for derslik in uygun_derslikler:
                            if derslik.kod not in kullanilan_derslik:  # Daha önce atanmadıysa
                                program[gun][saat] = (ders_kodu, ders_adi, derslik.kod)
                                derslik_doluluk[gun][saat].add(derslik.kod)
                                kullanilan_derslik.add(derslik.kod)

                                # Aynı bölümdeki aynı dersin her saatine aynı dersliği ata
                                derslik_atama[(bolum, ders_adi)] = derslik.kod

                                # Ortak bir dersse, diğer bölümlere de aynı dersliği ata
                                if ders_adi in ortak_d:
                                    ortak_derslik_atama[ders_adi] = derslik.kod
                                break  # Derslik atandıktan sonra döngüden çık

   
                                
# ---DERS PROGRAMI OLUŞTURMA İŞLEMLERİ---

    @staticmethod
    def dersleri_oku(donem, derslikler):
        bm_dersler = {}  # BM derslerini saklamak için sözlük
        ym_dersler = {}  # YM derslerini saklamak için sözlük

        try:
            # Veritabanından belirtilen dönemin BM ve YM derslerini çekiyoruz
            bm_dersler_l = session.query(Ders).join(Bolum).filter(Bolum.kod == "BM", Ders.donem == donem).all()
            ym_dersler_l = session.query(Ders).join(Bolum).filter(Bolum.kod == "YM", Ders.donem == donem).all()

            # BM derslerini sözlüğe yerleştiriyoruz
            for ders in bm_dersler_l:
                ders_bilgisi = {
                    "Dönem": ders.donem,
                    "Ders Adı": ders.ad,
                    "Ders Tipi": ders.ders_tipi,
                    "Teorik": ders.teorik_saat,
                    "Pratik": ders.uyg_saat
                }
                bm_dersler[ders.kod] = ders_bilgisi

            # YM derslerini sözlüğe yerleştiriyoruz
            for ders in ym_dersler_l:
                ders_bilgisi = {
                    "Dönem": ders.donem,
                    "Ders Adı": ders.ad,
                    "Ders Tipi": ders.ders_tipi,
                    "Teorik": ders.teorik_saat,
                    "Pratik": ders.uyg_saat
                }
                ym_dersler[ders.kod] = ders_bilgisi

            bm_program, ym_program = Program.ders_programi_olustur(bm_dersler, ym_dersler, derslikler)
            Program.programi_goster(bm_program, "BM")
            Program.programi_goster(ym_program, "YM")
            Program.excele_yazdir(bm_program, "BM", donem)
            Program.excele_yazdir(ym_program, "YM", donem)


        except Exception as e:
            print(f"Hata oluştu: {e}")

    @staticmethod
    def ders_programi_olustur(bm_dersler, ym_dersler, derslikler):
        """ Ders programını oluşturur ve derslikleri atar. """

        gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
        saatler = list(range(9, 17))  # 9:00 - 17:00 saatleri

        bm_program = {gun: {saat: None for saat in saatler} for gun in gunler}
        ym_program = {gun: {saat: None for saat in saatler} for gun in gunler}
        
        bm_ortak_dersler, ym_ortak_dersler, bm_ozel_dersler, ym_ozel_dersler = Program.dersleri_ayir(bm_dersler,
                                                                                                     ym_dersler)
        bm_lab_dersleri = Program.lab_dersleri_bul(bm_ozel_dersler, bm_ortak_dersler)
        ym_lab_dersleri = Program.lab_dersleri_bul(ym_ozel_dersler, ym_ortak_dersler)


        Program.ortak_dersleri_yerlestir(bm_ortak_dersler, ym_ortak_dersler, bm_program, ym_program)
        Program.dersi_yerlestir(bm_ozel_dersler, bm_program)
        Program.dersi_yerlestir(ym_ozel_dersler, ym_program)

        # Derslik ataması yap
        Program.derslikleri_atama(bm_program, ym_program, bm_dersler, ym_dersler, bm_ortak_dersler, derslikler)

        return bm_program, ym_program

    @staticmethod
    def dersleri_ayir(bm_dersler, ym_dersler):
        bm_ortak_dersler = {
            bm_kod: bm_ders
            for bm_kod, bm_ders in bm_dersler.items()
            for ym_kod, ym_ders in ym_dersler.items()
            if bm_ders["Ders Adı"] == ym_ders["Ders Adı"] and "lab" not in bm_ders["Ders Adı"].lower()
        }

        ym_ortak_dersler = {
            ym_kod: ym_ders
            for bm_kod, bm_ders in bm_dersler.items()
            for ym_kod, ym_ders in ym_dersler.items()
            if bm_ders["Ders Adı"] == ym_ders["Ders Adı"] and "lab" not in ym_ders["Ders Adı"].lower()
        }

        bm_ozel_dersler = {k: v for k, v in bm_dersler.items() if k not in bm_ortak_dersler}
        ym_ozel_dersler = {k: v for k, v in ym_dersler.items() if k not in ym_ortak_dersler}

        return bm_ortak_dersler, ym_ortak_dersler, bm_ozel_dersler, ym_ozel_dersler

    @staticmethod
    def lab_dersleri_bul(ozel_dersler, ortak_dersler):
        return {k: v for k, v in {**ozel_dersler, **ortak_dersler}.items() if "lab" in v["Ders Adı"].lower()}

    @staticmethod
    def uygun_saat_bul(hedef_program, gun, saat_sayisi):
        saatler = list(range(9, 17))  # Saat aralığı 9:00-17:00
        uygun_saatler = [saat for saat in saatler if hedef_program[gun][saat] is None]

        # Uygun saatler arasında 1 saatlik boşluk bırakılacak şekilde kontrol
        for i in range(len(uygun_saatler) - saat_sayisi):  # Boşluk eklemek için 1 saat daha ekleniyor
            if all(hedef_program[gun][saat] is None for saat in uygun_saatler[i:i + saat_sayisi]):
                return uygun_saatler[i]  # İlk uygun saat
        return None

    @staticmethod
    def dersi_yerlestir(dersler, hedef_program):
        ders_listesi = list(dersler.items())  # Derslerin listesini oluştur
        random.shuffle(ders_listesi)  # Dersleri karıştır

        for ders_kodu, ders in ders_listesi:
            saat_sayisi = ders["Teorik"] + ders["Pratik"]  # Dersin süre hesaplaması

            for gun in hedef_program:
                if len(set(hedef_program[gun].values()) - {None}) >= 2:  # İki ders eklenmişse
                    continue

                bosluklu_saat = Program.uygun_saat_bul(hedef_program, gun, saat_sayisi)
                if bosluklu_saat is None:
                    continue

                    # Dersin yerleştirilmesi
                for i in range(saat_sayisi):
                    hedef_program[gun][bosluklu_saat + i] = (ders_kodu, ders["Ders Adı"])

                bos_saat_sonraki = bosluklu_saat + saat_sayisi
                if bos_saat_sonraki < 17:
                    hedef_program[gun][bos_saat_sonraki] = ""

                break  # Ders yerleştirildikten sonra döngüden çıkıyoruz

    @staticmethod
    def ortak_dersleri_yerlestir(bm_ortak_dersler, ym_ortak_dersler, bm_program, ym_program):
        ortak_dersler = [
            (bm_kod, bm_ders, ym_kod, ym_ders)
            for bm_kod, bm_ders in bm_ortak_dersler.items()
            for ym_kod, ym_ders in ym_ortak_dersler.items()
            if bm_ders["Ders Adı"] == ym_ders["Ders Adı"]
        ]

        random.shuffle(ortak_dersler)

        for bm_kod, bm_ders, ym_kod, ym_ders in ortak_dersler:
            saat_sayisi = bm_ders["Teorik"] + bm_ders["Pratik"]

            for gun in bm_program:
                bos_saat = Program.uygun_saat_bul(bm_program, gun, saat_sayisi)

                if bos_saat:
                    # Dersin yerleştirilmesi
                    for i in range(saat_sayisi):
                        bm_program[gun][bos_saat + i] = (bm_kod, bm_ders["Ders Adı"])
                        ym_program[gun][bos_saat + i] = (ym_kod, ym_ders["Ders Adı"])

                    # **Boşluk ekleniyor**: Dersin bitişinden 1 saat sonrasına boşluk
                    bos_saat_sonraki = bos_saat + saat_sayisi
                    if bos_saat_sonraki < 17:  # Program sınırlarını aşmamak için kontrol
                        bm_program[gun][bos_saat_sonraki] = ""
                        ym_program[gun][bos_saat_sonraki] = ""

                    break

    @staticmethod
    def programi_goster(program, bolum_adi):
        """ Ders programını ekrana yazdırır (derslik bilgisi dahil). """
        print(f"\n{bolum_adi} Ders Programı:")
        for gun, saatler in program.items():
            print(f"\n{gun}:")
            for saat, ders in saatler.items():
                if ders:
                    ders_kodu, ders_adi = ders[:2]
                    derslik = ders[2] if len(ders) > 2 else "Bilinmiyor"
                    print(f"{saat}:00 - {ders_adi} ({ders_kodu}) - Derslik: {derslik}")

    # ---EXCELE YAZDIRMA İŞLEMLERİ---

    def excele_yazdir(program, bolum_adi, donem):
        dosya_adi = "ProgramŞablon.xlsx"  # Excel dosya adı

        try:
            # Excel dosyasını aç
            wb = openpyxl.load_workbook(dosya_adi)

            # Eğer sayfa zaten varsa uyarı ver, yoksa oluştur
            if bolum_adi in wb.sheetnames:
                print(f"⚠️ '{bolum_adi}' adlı sayfa zaten mevcut. Yeni sayfa oluşturulmadı.")
                Program.program_excele_yaz(wb, program, bolum_adi, donem)
                wb.save(dosya_adi)
            else:
                # Burada program içeriğini yazdırabilirsin, şu an boş sayfa oluşturuluyor
                Program.yeni_sayfa_olustur(wb, bolum_adi)
                wb.save(dosya_adi)
                Program.program_excele_yaz(wb, program, bolum_adi, donem)
                wb.save(dosya_adi)
                print(f"✅ '{bolum_adi}' sayfası başarıyla eklendi ve program yazıldı.")

            print(f"\n✅ '{bolum_adi}' için Program yazıldı.")

        except Exception as e:
            print(f"Hata oluştu: {e}")

    def yeni_sayfa_olustur(wb, bolum_adi):

        try:
            # İlk sayfayı al
            ilk_sayfa = wb.worksheets[0]

            # Yeni bir sayfa oluştur ve ilk sayfanın içeriğini kopyala
            yeni_sayfa = wb.create_sheet(title=bolum_adi)  # Yeni sheet adı bölüm adı olacak

            # Hücre genişliklerini kopyala
            for col in range(1, ilk_sayfa.max_column + 1):
                col_letter = get_column_letter(col)
                yeni_sayfa.column_dimensions[col_letter].width = ilk_sayfa.column_dimensions[col_letter].width

            # Hücre verileri ve formatları kopyala
            for row in ilk_sayfa.iter_rows():
                for cell in row:
                    # Yeni sayfada aynı hücreyi oluştur
                    yeni_hucre = yeni_sayfa[cell.coordinate]
                    yeni_hucre.value = cell.value  # Değeri kopyala

                    # Hücre stilini kopyala (font, renk, kenarlık, hizalama)
                    if cell.font:
                        yeni_hucre.font = openpyxl.styles.Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color
                        )
                    if cell.fill:
                        yeni_hucre.fill = openpyxl.styles.PatternFill(
                            fill_type=cell.fill.fill_type,
                            fgColor=cell.fill.fgColor,
                            bgColor=cell.fill.bgColor
                        )
                    if cell.border:
                        yeni_hucre.border = openpyxl.styles.Border(
                            left=cell.border.left,
                            right=cell.border.right,
                            top=cell.border.top,
                            bottom=cell.border.bottom,
                            diagonal=cell.border.diagonal,
                            diagonal_direction=cell.border.diagonal_direction,
                            outline=cell.border.outline,
                            vertical=cell.border.vertical,
                            horizontal=cell.border.horizontal
                        )
                    if cell.alignment:
                        yeni_hucre.alignment = openpyxl.styles.Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent
                        )
                    if cell.number_format:
                        yeni_hucre.number_format = cell.number_format

            yeni_sayfa["C1"] = bolum_adi


        except Exception as e:
            print(f" Hata oluştu: {e}")

    def program_excele_yaz(wb, program, bolum_adi, donem):
        try:
            satir = 4  # Başlangıç satırı
            sayfa = wb[bolum_adi]
            sutun = "C" if int(donem) in (1, 2) else \
                    "D" if int(donem) in (3, 4) else \
                    "E" if int(donem) in (5, 6) else \
                    "F" if int(donem) in (7, 8) else None

            if not sutun:
                print(f"⚠️ Geçersiz dönem: {donem}")
                return

            for gun, dersler in program.items():
                for saat, ders_bilgisi in dersler.items():
                    hucre = f"{sutun}{satir}"
                    
                    if ders_bilgisi and ders_bilgisi != "":
                        ders_kodu, ders_adi, sinif = ders_bilgisi
                        kod = ders_kodu[3:]
                        
                        # Öğretim üyesini bul
                        ders = session.query(Ders).filter_by(kod=ders_kodu).first()
                        ogretim_uyesi_adi = ders.ogretim_uyesi.ad if ders and ders.ogretim_uyesi else "Belirtilmemiş"
                        
                        sayfa[hucre] = f"{kod} - {ders_adi}\n{ogretim_uyesi_adi} ({sinif})\n"
                    else:
                        sayfa[hucre] = "\n "

                    if satir == 43:
                        continue
                    else:
                        satir += 1

            print(f"✅ '{bolum_adi}' için program başarıyla Excel'e yazıldı.")

        except Exception as e:
            print(f"Programı yazarken hata oluştu: {e}")

# Menü çalıştırılıyor
Sistem.menu()